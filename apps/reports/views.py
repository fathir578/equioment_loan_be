import os
from datetime import datetime
from urllib import request
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from core.permissions import IsAdminOrPetugas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from apps.departments.models import Department
from apps.loans.models import Loan, LoanItem
from apps.returns.models import Return


class LaporanJurusanView(APIView):
    """
    GET /api/v1/reports/{dept_kode}/
    Download laporan peminjaman per jurusan dalam format Excel.
    """
    permission_classes = [IsAdminOrPetugas]

    def get(self, request, dept_kode):
        department = get_object_or_404(Department, kode=dept_kode.upper())
        if request.user.role == 'petugas':
            if request.user.department != department:
                return Response(
                {
                    "success": False,
                    "message": "Anda hanya bisa akses laporan jurusan Anda sendiri."
                },
            status=403
        )
        
        # Query data peminjaman terkait jurusan ini
        # Ambil LoanItem yang user-nya di jurusan ini ATAU alat-nya milik jurusan ini
        # Namun instruksi "Laporan Jurusan" biasanya mengacu pada peminjam dari jurusan tersebut
        loan_items = LoanItem.objects.filter(
            loan__user__department=department
        ).select_related('loan__user', 'tool', 'loan').order_by('-loan__loan_date')

        # Buat Workbook
        wb = Workbook()
        
        # 1. Sheet "Semua Kelas"
        ws_all = wb.active
        ws_all.title = "Semua Kelas"
        self._write_sheet_data(ws_all, loan_items, f"LAPORAN PEMINJAMAN - JURUSAN {department.nama}")

        # 2. Sheet per Tingkatan (X, XI, XII)
        # Kelompokkan data berdasarkan tingkatan
        tingkatan_map = {}
        for item in loan_items:
            kelas = item.loan.user.kelas or "TANPA KELAS"
            # Extract tingkatan: "X RPL A" -> "X RPL"
            parts = kelas.split()
            tingkatan = ' '.join(parts[:2]) if len(parts) >= 2 else parts[0]
            
            if tingkatan not in tingkatan_map:
                tingkatan_map[tingkatan] = []
            tingkatan_map[tingkatan].append(item)

        # Buat sheet untuk setiap tingkatan
        for tingkatan, items in sorted(tingkatan_map.items()):
            ws = wb.create_sheet(title=tingkatan)
            self._write_sheet_data(ws, items, f"LAPORAN PEMINJAMAN - {tingkatan}")

        # Simpan file ke folder reports/{DEPT}/
        now = datetime.now()
        timestamp = now.strftime("%Y_%m")
        filename = f"laporan_{department.kode}_{timestamp}.xlsx"
        
        relative_folder = os.path.join('reports', department.kode)
        full_folder = os.path.join(settings.MEDIA_ROOT, relative_folder)
        os.makedirs(full_folder, exist_ok=True)
        
        file_path = os.path.join(full_folder, filename)
        wb.save(file_path)

        # Response Download
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response

    def _write_sheet_data(self, ws, items, title):
        # Styling
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Header Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.cell(row=1, column=1).value = title
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = center_align

        # Table Header
        headers = [
            'No', 'NIS', 'Nama', 'Kelas', 'Alat', 'Jumlah', 
            'Tgl Pinjam', 'Tgl Kembali', 'Status', 'Denda'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Data Rows
        for i, item in enumerate(items, 1):
            row_idx = i + 3
            user = item.loan.user
            
            # Cari tgl kembali terakhir (jika ada)
            returns = Return.objects.filter(loan=item.loan).order_by('-return_date')
            tgl_kembali = returns.first().return_date.strftime("%d/%m/%Y") if returns.exists() else "-"
            
            # Total denda loan
            total_denda = sum(r.total_fine for r in returns) if returns.exists() else 0

            data = [
                i,
                user.nis or "-",
                user.nama_lengkap or user.username,
                user.kelas or "-",
                item.tool.name,
                item.quantity,
                item.loan.loan_date.strftime("%d/%m/%Y"),
                tgl_kembali,
                item.loan.get_status_display(),
                f"Rp{total_denda:,.0f}"
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = val
                cell.border = border
                if col in [1, 2, 4, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal='center')

        # Auto-width (Start from row 3 to avoid merged title cell)
        for col_idx in range(1, 11):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_idx).column_letter
            # Check headers (row 3) and data (row 4 onwards)
            for row in ws.iter_rows(min_row=3, max_col=10):
                cell = row[col_idx-1]
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 3)
            ws.column_dimensions[column_letter].width = adjusted_width
